from __future__ import annotations

import json
import subprocess
import shutil
import tempfile
import zipfile
from io import BytesIO
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "약제팀 라벨" / "원내보유의약품리스트.xlsx"
DATA = ROOT / "약제팀 라벨" / "data" / "hospitalDrugLabels.generated.json"
SHEET_XML = "xl/worksheets/sheet1.xml"
NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
FLUID_COLOR_HEADER = "일반수액 색기호"
BORDER_COLOR_HEADER = "테두리 색기호"
IMAGE_PATH_HEADER = "식별사진경로"
IMAGE_SOURCE_HEADER = "식별사진출처"
ET.register_namespace("", NS)
ET.register_namespace("mc", MC_NS)


def inline_cell(reference: str, value: str) -> ET.Element:
    cell = ET.Element(f"{{{NS}}}c", {"r": reference, "t": "inlineStr"})
    inline = ET.SubElement(cell, f"{{{NS}}}is")
    text = ET.SubElement(inline, f"{{{NS}}}t")
    text.text = value
    return cell


def column_values(workbook, header: str) -> dict[str, str]:
    worksheet = workbook.worksheets[0]
    headers = [str(value or "").strip() for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "약품코드" not in headers or header not in headers:
        return {}
    code_index = headers.index("약품코드")
    value_index = headers.index(header)
    return {
        str(row[code_index] or "").strip(): str(row[value_index] or "").strip()
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if code_index < len(row) and str(row[code_index] or "").strip()
    }


def main() -> None:
    rows = json.loads(DATA.read_text(encoding="utf-8"))
    image_by_code = {
        str(row["code"]): (
            str(row.get("imagePath", "")).lstrip("/"),
            str(row.get("imageSourceUrl", "")),
        )
        for row in rows
    }

    base_blob = subprocess.run(
        ["git", "show", f"HEAD:{SOURCE.relative_to(ROOT).as_posix()}"],
        cwd=ROOT,
        check=True,
        capture_output=True,
    ).stdout
    base_workbook = load_workbook(BytesIO(base_blob), read_only=True, data_only=True)
    base_fluid_colors = column_values(base_workbook, FLUID_COLOR_HEADER)
    base_border_colors = column_values(base_workbook, BORDER_COLOR_HEADER)
    base_workbook.close()

    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]
    max_row = worksheet.max_row
    row_values: dict[int, tuple[str, str, str, str]] = {
        1: (FLUID_COLOR_HEADER, BORDER_COLOR_HEADER, IMAGE_PATH_HEADER, IMAGE_SOURCE_HEADER),
    }
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True), start=2):
        code = str(row[0] or "").strip()
        image_path, source_url = image_by_code.get(code, ("", ""))
        row_values[row_number] = (base_fluid_colors.get(code, ""), base_border_colors.get(code, ""), image_path, source_url)
    workbook.close()

    with zipfile.ZipFile(SOURCE, "r") as source_zip:
        root = ET.fromstring(source_zip.read(SHEET_XML))
        dimension = root.find(f"{{{NS}}}dimension")
        if dimension is not None:
            start = dimension.attrib.get("ref", "A1").split(":")[0]
            dimension.set("ref", f"{start}:AQ{max_row}")

        sheet_data = root.find(f"{{{NS}}}sheetData")
        if sheet_data is None:
            raise RuntimeError("약품조회 시트의 sheetData를 찾을 수 없습니다.")
        for row_element in sheet_data.findall(f"{{{NS}}}row"):
            row_number = int(row_element.attrib["r"])
            if row_number not in row_values:
                continue
            row_element.set("spans", "1:43")
            for cell in list(row_element):
                if cell.attrib.get("r", "").startswith(("AN", "AO", "AP", "AQ")):
                    row_element.remove(cell)
            fluid_color, border_color, image_path, source_url = row_values[row_number]
            row_element.append(inline_cell(f"AN{row_number}", fluid_color))
            row_element.append(inline_cell(f"AO{row_number}", border_color))
            row_element.append(inline_cell(f"AP{row_number}", image_path))
            row_element.append(inline_cell(f"AQ{row_number}", source_url))

        patched_xml = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        if b'Ignorable="x14ac"' in patched_xml and b"xmlns:x14ac=" not in patched_xml:
            patched_xml = patched_xml.replace(
                b"<worksheet ",
                f'<worksheet xmlns:x14ac="{X14AC_NS}" '.encode("utf-8"),
                1,
            )
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=SOURCE.parent) as temporary:
            temporary_path = Path(temporary.name)
        try:
            with zipfile.ZipFile(temporary_path, "w") as target_zip:
                for item in source_zip.infolist():
                    target_zip.writestr(item, patched_xml if item.filename == SHEET_XML else source_zip.read(item.filename))
            shutil.copy2(temporary_path, SOURCE)
        finally:
            temporary_path.unlink(missing_ok=True)

    print(f"Patched image columns for {len(row_values) - 1} rows: {SOURCE}")


if __name__ == "__main__":
    main()
