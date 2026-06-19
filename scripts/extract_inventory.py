from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "src" / "data" / "inventory.generated.json"


SHEET_ALIAS = {
    "42W": "42",
    "61W": "61",
    "62W": "62",
    "71W": "71",
    "72W": "72",
    "81W": "81",
    "82W": "82",
    "91W": "91",
    "92W": "92",
    "101W": "101",
    "102W": "102",
    "111W": "111",
    "112W": "112",
    "121W": "121",
    "신속대응팀": "RRT",
    "HBEF심혈관조영실": "HBEF",
    "체외순환실": "HEART",
    "외래주사실": "INJ",
    "영상의학과": "DRO",
    "안과": "OT",
    "재활의학과": "RH",
    "산부인과": "OG",
    "난임클리닉": "난임",
    "이비인후과": "OL",
    "피부과": "DM",
    "소화기병검사실": "GICLA",
    "정형외과": "OS",
    "신경과": "NR",
    "비뇨기과": "UR",
}

STOCK_CODE_OVERRIDES = {
    "0.9% NaKCl 20mEq/100ml btl": "XNAK20",
}

CHECKLIST_LABEL_ROWS = {"양호불량"}

STOCK_FIELD_OVERRIDES = {
    "XBPCA5W": {"warning": ""},
    "XEPIN": {"storageType": "ROOM"},
    "XKPHMB": {"warning": "고위험의약품"},
    "XMEXO": {"warning": "유사모양"},
    "XMVH": {"storageType": "REFRIGERATED"},
    "XNA40": {"warning": "고위험의약품"},
}

ECART_FIELD_OVERRIDES = {
    "XADENO6": {"name": "Adenocor( Adenosin )", "dosage": "6mg/Vial", "quantity": 3},
    "XLID2W": {"name": "2% Lidocaine 400mg", "dosage": "2% 20mL/Vial", "quantity": 2},
    "XNB84": {"name": "Sodium Bicabonate", "dosage": "20mEq/20mL/Amp", "quantity": 10},
    "XNITR10F": {"name": "Nitrolingual 0.1%", "dosage": "10mg/10ml", "quantity": 5},
    "XNS20": {"name": "N/S 20cc", "dosage": "20mL/Amp", "quantity": 3},
    "XCPENIR": {"name": "Peniramin", "dosage": "4mg/2ml/Amp", "quantity": 3},
    "NITR": {"name": "Nitroglycerin(SL)", "dosage": "0.6mg/Tab", "quantity": 3},
}


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return "" if text in {"#N/A", "None"} else re.sub(r"\s+", " ", text)


def qty(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    text = clean(value)
    if not text:
        return 0
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return int(float(match.group(0))) if match else 0


def room_update_date(value: Any) -> str:
    text = clean(value).replace(" ", "")
    match = re.fullmatch(r"(\d{2,4})\.(\d{1,2})\.?(\d{1,2})", text)
    if not match:
        return ""
    year, month, day = match.groups()
    return f"{year[-2:]}.{int(month):02d}.{int(day):02d}"


def collect_room_update_dates(wb: Any) -> dict[str, str]:
    dates: dict[str, str] = {}
    for ws in wb.worksheets[1:]:
        for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
            for value in row:
                date = room_update_date(value)
                if date:
                    dates[ws.title] = date
                    break
            if ws.title in dates:
                break
    return dates


def storage_type(storage: str) -> str:
    text = storage.replace(" ", "")
    if "냉장보관하지" in text:
        return "ROOM"
    if any(token in text for token in ["냉장", "2~8", "2-8", "2∼8", "5±3", "5℃이하", "10℃이하"]):
        return "REFRIGERATED"
    if "차광" in text:
        return "LIGHT_PROTECTED"
    return "ROOM"


def stock_code(raw_code: str, product_name: str) -> str:
    return STOCK_CODE_OVERRIDES.get(product_name, raw_code)


def is_checklist_label_row(text: str) -> bool:
    return re.sub(r"\s+", "", text) in CHECKLIST_LABEL_ROWS


def find_workbook(keyword: str) -> Path:
    matches = [p for p in ROOT.glob("*.xlsx") if keyword in p.name and not p.name.startswith("~$")]
    if not matches:
        raise FileNotFoundError(f"Cannot find workbook containing {keyword!r}")
    return matches[0]


def collect_warnings(stock_path: Path) -> dict[str, str]:
    wb = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
    warnings: dict[str, set[str]] = defaultdict(set)
    for ws in wb.worksheets[1:]:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [clean(v) for v in rows[0]]
        if "약품코드" not in header or "혼돈주의약품" not in header:
            continue
        code_idx = header.index("약품코드")
        warning_idx = header.index("혼돈주의약품")
        for row in rows[1:]:
            if code_idx >= len(row) or warning_idx >= len(row):
                continue
            code = clean(row[code_idx])
            warning = clean(row[warning_idx])
            if code and warning:
                warnings[code].add(warning)
    return {code: ", ".join(sorted(values)) for code, values in warnings.items()}


def parse_stock(stock_path: Path) -> dict[str, Any]:
    warnings = collect_warnings(stock_path)
    wb = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
    room_update_dates = collect_room_update_dates(wb)
    ws = wb.worksheets[0]
    header = [clean(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    room_columns = [(idx, name) for idx, name in enumerate(header[6:], start=6) if name and name != "합계"]

    drugs = []
    allocations = []
    room_stats = {name: {"allocationCount": 0, "totalQuantity": 0} for _, name in room_columns}

    for row in ws.iter_rows(min_row=2, values_only=True):
        raw_code = clean(row[0] if len(row) > 0 else "")
        product_name = clean(row[2] if len(row) > 2 else "")
        code = stock_code(raw_code, product_name)
        if not raw_code or not code:
            continue
        drug = {
            "code": code,
            "genericName": clean(row[1] if len(row) > 1 else ""),
            "productName": product_name,
            "spec": clean(row[3] if len(row) > 3 else ""),
            "storage": clean(row[4] if len(row) > 4 else ""),
            "note": clean(row[5] if len(row) > 5 else ""),
            "warning": warnings.get(code, warnings.get(raw_code, clean(row[5] if len(row) > 5 else ""))),
            "storageType": storage_type(clean(row[4] if len(row) > 4 else "")),
        }
        drug.update(STOCK_FIELD_OVERRIDES.get(code, {}))
        drugs.append(drug)
        for idx, room_name in room_columns:
            required_qty = qty(row[idx] if idx < len(row) else None)
            if required_qty > 0:
                allocations.append({"roomId": room_name, "drugCode": code, "requiredQty": required_qty})
                room_stats[room_name]["allocationCount"] += 1
                room_stats[room_name]["totalQuantity"] += required_qty

    rooms = [
        {
            "id": name,
            "label": name,
            "sourceColumn": name,
            "sourceSheet": SHEET_ALIAS.get(name, name),
            "sourceUpdatedAt": room_update_dates.get(SHEET_ALIAS.get(name, name), ""),
            **room_stats[name],
        }
        for _, name in room_columns
    ]
    return {"drugs": drugs, "rooms": rooms, "allocations": allocations}


def parse_ecart(ecart_path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(ecart_path, read_only=True, data_only=True)
    general_ws = wb.worksheets[0]
    general_items = []
    for row in general_ws.iter_rows(min_row=1, values_only=True):
        if not clean(row[0] if len(row) > 0 else "").isdigit():
            continue
        code = clean(row[1] if len(row) > 1 else "")
        name = clean(row[2] if len(row) > 2 else "")
        if not code or not name:
            continue
        item = {
            "id": code,
            "code": code,
            "name": name,
            "dosage": clean(row[3] if len(row) > 3 else ""),
            "quantity": qty(row[4] if len(row) > 4 else None),
        }
        item.update(ECART_FIELD_OVERRIDES.get(code, {}))
        general_items.append(item)

    nicu_items = []
    nicu_ws = wb.worksheets[1]
    for row in nicu_ws.iter_rows(min_row=1, values_only=True):
        no = clean(row[0] if len(row) > 0 else "")
        if not no.isdigit():
            continue
        name = clean(row[1] if len(row) > 1 else "")
        if not name:
            continue
        nicu_items.append(
            {
                "id": f"NICU-{int(no):02d}",
                "code": "",
                "name": name,
                "dosage": clean(row[2] if len(row) > 2 else ""),
                "quantity": qty(row[4] if len(row) > 4 else None),
            }
        )

    departments = []
    receive_ws = wb.worksheets[2]
    for row in receive_ws.iter_rows(min_row=3, values_only=True):
        value = clean(row[0] if len(row) > 0 else "")
        if value:
            departments.append(value)
    return {"generalItems": general_items, "nicuItems": nicu_items, "departments": departments}


def parse_checklist(checklist_path: Path) -> list[dict[str, str]]:
    wb = openpyxl.load_workbook(checklist_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    items = []
    current_section = ""
    for row in ws.iter_rows(values_only=True):
        cells = [clean(value) for value in row]
        nonempty = [cell for cell in cells if cell and cell != "□"]
        if not nonempty:
            continue
        first = nonempty[0]
        if first.startswith("[") and first.endswith("]"):
            current_section = first.strip("[]")
            rest = " ".join(nonempty[1:])
            if rest:
                items.append({"section": current_section, "text": rest})
            continue
        if "병동 비품약&E-cart 점검 체크리스트" in first or "점검 내용을" in first:
            continue
        text = " ".join(nonempty)
        if text and current_section and not is_checklist_label_row(text):
            if text.startswith("2-1 ") and " 2-2 " in text:
                first, second = text.split(" 2-2 ", 1)
                items.append({"section": current_section, "text": first})
                items.append({"section": current_section, "text": f"2-2 {second}"})
            else:
                items.append({"section": current_section, "text": text})
    if not any(item["section"] == "냉장약" and item["text"].startswith("6.") for item in items):
        insert_at = max((idx for idx, item in enumerate(items) if item["section"] == "냉장약"), default=-1) + 1
        items.insert(insert_at, {"section": "냉장약", "text": "6. 연 1회 냉장고 온도계 검증 여부"})
    return items


def main() -> None:
    stock_path = find_workbook("202606")
    ecart_path = find_workbook("E-cart")
    checklist_path = find_workbook("체크리스트")
    stock = parse_stock(stock_path)
    ecart = parse_ecart(ecart_path)
    checklist = parse_checklist(checklist_path)
    data = {
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "sourceFiles": {
            "stockWorkbook": stock_path.name,
            "ecartWorkbook": ecart_path.name,
            "checklistWorkbook": checklist_path.name,
        },
        "summary": {
            "stockDrugCount": len(stock["drugs"]),
            "stockRoomCount": len(stock["rooms"]),
            "stockAllocationCount": len(stock["allocations"]),
            "ecartGeneralItemCount": len(ecart["generalItems"]),
            "ecartNicuItemCount": len(ecart["nicuItems"]),
            "ecartDepartmentCount": len(ecart["departments"]),
            "checklistItemCount": len(checklist),
        },
        "stock": stock,
        "ecart": ecart,
        "checklist": checklist,
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(data["summary"], ensure_ascii=False, indent=2))
    print(f"wrote {OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
