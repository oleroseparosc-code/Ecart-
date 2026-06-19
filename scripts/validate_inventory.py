from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "src" / "data" / "inventory.generated.json"


def clean(value):
    return "" if value is None else str(value).strip()


def qty(value):
    if value is None or clean(value) == "":
        return 0
    match = re.search(r"-?\d+(?:\.\d+)?", clean(value))
    return int(float(match.group(0))) if match else 0


def find_workbook(keyword: str) -> Path:
    return [p for p in ROOT.glob("*.xlsx") if keyword in p.name and not p.name.startswith("~$")][0]


def main() -> None:
    data = json.loads(DATA.read_text(encoding="utf-8"))
    stock_path = find_workbook("202606")
    wb = openpyxl.load_workbook(stock_path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    header = [clean(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rooms = [h for h in header[6:] if h and h != "합계"]
    rows = [row for row in ws.iter_rows(min_row=2, values_only=True) if clean(row[0])]
    expected_allocations = 0
    samples = {
        ("42W", "XAPH1"): 2,
        ("71W", "XAMIOD"): 4,
        ("PED", "XXTDAP5W"): 5,
        ("외래주사실", "XXARIPIP3"): 3,
    }
    observed_samples = {}
    for row in rows:
        code = clean(row[0])
        for idx, room in enumerate(rooms, start=6):
            count = qty(row[idx] if idx < len(row) else None)
            if count > 0:
                expected_allocations += 1
            if (room, code) in samples:
                observed_samples[(room, code)] = count

    assert data["summary"]["stockDrugCount"] == len(rows), "stock drug count mismatch"
    assert data["summary"]["stockRoomCount"] == len(rooms), "stock room count mismatch"
    assert data["summary"]["stockAllocationCount"] == expected_allocations, "allocation count mismatch"
    for key, expected in samples.items():
        actual = observed_samples.get(key)
        assert actual == expected, f"sample {key} expected {expected}, got {actual}"

    by_key = {(a["roomId"], a["drugCode"]): a["requiredQty"] for a in data["stock"]["allocations"]}
    for key, expected in samples.items():
        assert by_key.get(key) == expected, f"generated sample {key} expected {expected}"

    ecart_path = find_workbook("E-cart")
    ewb = openpyxl.load_workbook(ecart_path, read_only=True, data_only=True)
    general_rows = [
        row
        for row in ewb.worksheets[0].iter_rows(values_only=True)
        if clean(row[0]).isdigit() and clean(row[1]) and clean(row[2])
    ]
    assert data["summary"]["ecartGeneralItemCount"] == len(general_rows), "E-cart item count mismatch"
    print("inventory validation passed")


if __name__ == "__main__":
    main()
