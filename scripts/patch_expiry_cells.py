from __future__ import annotations

import os
import re
import tempfile
import zipfile
from collections import defaultdict
from datetime import date
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "약제팀 라벨" / "원내보유의약품리스트.xlsx"
SALES_WORKBOOK = Path("G:/동국대학교일산병원_매출_날짜_20260813.xlsx")
MANUAL_EXPIRY_BY_DRUG_CODE = {"DGX": "2027-05-26"}
MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL = "http://schemas.openxmlformats.org/package/2006/relationships"


def tag(name: str) -> str:
    return f"{{{MAIN}}}{name}"


def column_letters(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def date_key(value: object) -> str:
    if isinstance(value, date):
        return value.isoformat()
    match = re.search(r"(20\d{2})[./-]?(\d{1,2})[./-]?(\d{1,2})", str(value or ""))
    return f"{match[1]}-{match[2].zfill(2)}-{match[3].zfill(2)}" if match else ""


def excel_serial(value: str) -> int:
    return (date.fromisoformat(value) - date(1899, 12, 30)).days


def normalized_name(value: object) -> str:
    text = str(value or "").lower().replace("캅셀", "캡슐")
    text = re.sub(r"\bnew\b|\(new\)|서방|필름코팅|장용|연질|캡슐|정|주", "", text)
    text = re.sub(r"(?<=\d)(?:mcg|mg|ml|iu|g|%|t|c)\b", "", text)
    text = re.sub(r"\b(?:tab|cap|inj|syr|btl|box|pkg|vial|amp)\b", "", text)
    return re.sub(r"[^0-9a-z가-힣]+", "", text)


def number_tokens(value: object) -> set[str]:
    return set(re.findall(r"(?<![0-9])\d+(?:\.\d+)?(?![0-9])", str(value or "")))


def contains_numbers(value: str, required: set[str]) -> bool:
    return all(re.search(rf"(?<![0-9]){re.escape(number)}(?![0-9])", value) for number in required)


def sales_matches() -> dict[str, tuple[str, str]]:
    workbook = load_workbook(SALES_WORKBOOK, data_only=True, read_only=True)
    sheet = workbook.active
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    name_index = headers.index("제 품 명")
    item_index = headers.index("물품코드")
    expiry_index = headers.index("유효기한")
    by_item_code: dict[str, list[tuple[str, str, str]]] = defaultdict(list)
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item_code = str(row[item_index] or "").strip()
        product_name = str(row[name_index] or "").strip()
        expiry = date_key(row[expiry_index])
        if item_code and product_name and expiry:
            by_item_code[item_code].append((normalized_name(product_name), product_name, expiry))
    workbook.close()

    source_rows = load_workbook(WORKBOOK, data_only=True, read_only=True)
    sheet = source_rows["약품조회"]
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    code_index = headers.index("약품코드")
    name_index = headers.index("한글약품명")
    result: dict[str, tuple[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = str(row[code_index] or "").strip()
        korean_name = str(row[name_index] or "").strip()
        normalized = normalized_name(korean_name)
        if not code or len(normalized) < 5:
            continue
        required_numbers = number_tokens(korean_name)
        candidates: list[tuple[str, list[tuple[str, str, str]]]] = []
        for item_code, sales_rows in by_item_code.items():
            if any(
                (sales_name.startswith(normalized) or normalized.startswith(sales_name))
                and contains_numbers(product_name, required_numbers)
                for sales_name, product_name, _ in sales_rows
            ):
                candidates.append((item_code, sales_rows))
        if len(candidates) != 1:
            continue
        item_code, sales_rows = candidates[0]
        expiry = min(expiry for _, _, expiry in sales_rows)
        result[code] = (item_code, expiry)
    source_rows.close()
    return result


def main() -> None:
    matches = sales_matches()
    workbook = load_workbook(WORKBOOK, data_only=True, read_only=True)
    sheet = workbook["약품조회"]
    headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    code_col = headers.index("약품코드") + 1
    item_col = headers.index("물품코드") + 1
    latest_col = headers.index("최신 유효기간") + 1
    updates: list[tuple[int, str, str]] = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        drug_code = str(row[code_col - 1] or "").strip()
        matched = matches.get(drug_code)
        if drug_code in MANUAL_EXPIRY_BY_DRUG_CODE:
            matched = (str(row[item_col - 1] or "").strip(), MANUAL_EXPIRY_BY_DRUG_CODE[drug_code])
        if not matched:
            continue
        item_code, expiry = matched
        current_item_code = str(row[item_col - 1] or "").strip()
        current_expiry = date_key(row[latest_col - 1])
        if current_item_code != item_code or current_expiry != expiry:
            updates.append((row_index, item_code, expiry))
    workbook.close()
    if not updates:
        print("No item-code or expiry updates required.")
        return

    with zipfile.ZipFile(WORKBOOK, "r") as source:
        workbook_xml = ET.fromstring(source.read("xl/workbook.xml"))
        relations = ET.fromstring(source.read("xl/_rels/workbook.xml.rels"))
        targets = {relation.attrib["Id"]: relation.attrib["Target"] for relation in relations.findall(f"{{{PACKAGE_REL}}}Relationship")}
        sheet_id = next(sheet.attrib[f"{{{REL}}}id"] for sheet in workbook_xml.findall(f".//{tag('sheet')}") if sheet.attrib["name"] == "약품조회")
        sheet_xml_path = "xl/" + targets[sheet_id].lstrip("/")
        sheet_root = ET.fromstring(source.read(sheet_xml_path))
        rows_by_index = {int(row.attrib["r"]): row for row in sheet_root.findall(f".//{tag('row')}")}
        latest_letter = column_letters(latest_col)
        date_style = next((cell.attrib.get("s") for cell in sheet_root.findall(f".//{tag('c')}") if cell.attrib.get("r", "").startswith(latest_letter) and cell.attrib.get("s")), None)

        def write_cell(row_xml: ET.Element, address: str, value: str, style: str | None = None, text: bool = False) -> None:
            cell = next((candidate for candidate in row_xml.findall(tag("c")) if candidate.attrib.get("r") == address), None)
            if cell is None:
                cell = ET.SubElement(row_xml, tag("c"), {"r": address})
            if text:
                cell.set("t", "str")
            else:
                cell.attrib.pop("t", None)
            if style:
                cell.set("s", style)
            for child in list(cell):
                cell.remove(child)
            ET.SubElement(cell, tag("v")).text = value

        for row_index, item_code, expiry in updates:
            row_xml = rows_by_index[row_index]
            write_cell(row_xml, f"{column_letters(item_col)}{row_index}", item_code, text=True)
            write_cell(row_xml, f"{latest_letter}{row_index}", str(excel_serial(expiry)), date_style)

        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=WORKBOOK.parent) as temp_file:
            temporary = Path(temp_file.name)
        with zipfile.ZipFile(temporary, "w", zipfile.ZIP_DEFLATED) as target:
            for entry in source.infolist():
                target.writestr(entry, ET.tostring(sheet_root, encoding="utf-8", xml_declaration=True) if entry.filename == sheet_xml_path else source.read(entry.filename))
    os.replace(temporary, WORKBOOK)
    print(f"Patched {len(updates)} sales-file item-code and expiry rows while preserving workbook formatting.")


if __name__ == "__main__":
    main()
